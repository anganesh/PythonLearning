import cx_Oracle, json
import db_config
from openpyxl import Workbook
# Opening JSON file
f = open('sqllist.json',)
  
# returns JSON object as 
# a dictionary
data = json.load(f)
sheetnames = list(data.keys())
print(sheetnames)
# Closing file
f.close()

con = cx_Oracle.connect(db_config.user, db_config.pw, db_config.dsn)

book= Workbook(write_only = True)

# Iterating through the json
# list
for i in data['sql_details']:
    vsheetname = i["sheetName"]
    print(vsheetname)
    vsql = i["sql"]
    print(vsql)

    rowheader = []
    rowdata = []
    sheet = book.create_sheet(vsheetname)



    cur = con.cursor()

    cur.execute(vsql)

    """ below cur.description to get the column header"""

    for column in cur.description:
        rowheader.append(column[0])
    sheet.append(rowheader)
    print(rowheader)


    """ appending sql data to each row"""
    
    res = cur.fetchall()
    for row in res:
        rowdata.append(row)
        sheet.append(row)
    print(rowdata)
    cur.close()

book.save("multisqla.xlsx")



con.close()  

